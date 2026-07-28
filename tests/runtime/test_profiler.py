from __future__ import annotations

import io
import json
import unittest
from datetime import UTC, datetime

from openpyxl import Workbook

from agent.runtime.profiler import profile_dataset_version
from application.datasets.models import DatasetVersion


class InMemoryStorage:
    def __init__(self) -> None:
        self.objects: dict[str, bytes] = {}

    def open_read(self, storage_key: str):
        if storage_key not in self.objects:
            raise FileNotFoundError(storage_key)
        return io.BytesIO(self.objects[storage_key])


class DatasetVersionProfilerTests(unittest.TestCase):
    def test_profiles_xlsx_with_multiple_sheets(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sales"
        sheet.append(["date", "sales"])
        sheet.append(["2026-07-01", 10])
        second = workbook.create_sheet("Stocks")
        second.append(["sku", "stock"])
        second.append(["A-1", 5])
        payload = io.BytesIO()
        workbook.save(payload)

        storage = InMemoryStorage()
        storage.objects["workspace/.blobs/xlsx"] = payload.getvalue()
        version = DatasetVersion(
            id="dsv_xlsx",
            dataset_id="ds_1",
            original_filename="sales.xlsx",
            storage_key="workspace/.blobs/xlsx",
            format="xlsx",
            size_bytes=len(storage.objects["workspace/.blobs/xlsx"]),
            checksum_sha256="sum-xlsx",
            status="profiling",
            created_at=DatasetVersion.model_fields["created_at"].annotation.now(),  # type: ignore[attr-defined]
        )

        result = profile_dataset_version(version, storage)

        self.assertTrue(result.success)
        self.assertEqual([sheet.name for sheet in result.profile.sheets], ["Sales", "Stocks"])
        self.assertEqual(result.profile.sheets[0].columns[0].name, "date")
        self.assertEqual(result.profile.sheets[0].row_count, 1)
        self.assertEqual(result.profile.sheets[1].row_count, 1)

    def test_profiles_utf8_bom_csv(self) -> None:
        storage = InMemoryStorage()
        storage.objects["workspace/.blobs/csv"] = (
            "\ufeffdate,sales\n2026-07-01,10\n2026-07-02,11\n".encode("utf-8")
        )
        version = _build_version("dsv_csv", "csv", "workspace/.blobs/csv")

        result = profile_dataset_version(version, storage)

        self.assertTrue(result.success)
        self.assertEqual(result.profile.sheets[0].columns[0].name, "date")
        self.assertEqual(result.profile.sheets[0].row_count, 2)

    def test_profiles_cp1251_csv_with_warning(self) -> None:
        storage = InMemoryStorage()
        storage.objects["workspace/.blobs/csv1251"] = (
            "дата;продажи\n2026-07-01;10\n".encode("cp1251")
        )
        version = _build_version("dsv_csv_1251", "csv", "workspace/.blobs/csv1251")

        result = profile_dataset_version(version, storage)

        self.assertTrue(result.success)
        self.assertTrue(any("cp1251" in warning for warning in result.profile.warnings))
        self.assertEqual(result.profile.sheets[0].columns[0].name, "дата")

    def test_profiles_json_list_of_objects(self) -> None:
        storage = InMemoryStorage()
        storage.objects["workspace/.blobs/json"] = json.dumps(
            [{"date": "2026-07-01", "sales": 10}, {"date": "2026-07-02", "sales": 11}]
        ).encode("utf-8")
        version = _build_version("dsv_json", "json", "workspace/.blobs/json")

        result = profile_dataset_version(version, storage)

        self.assertTrue(result.success)
        self.assertEqual(result.profile.sheets[0].columns[0].name, "date")
        self.assertEqual(result.profile.sheets[0].row_count, 2)

    def test_csv_header_only_is_invalid(self) -> None:
        storage = InMemoryStorage()
        storage.objects["workspace/.blobs/header-only"] = b"date,sales\n"
        version = _build_version("dsv_header_only", "csv", "workspace/.blobs/header-only")

        result = profile_dataset_version(version, storage)

        self.assertFalse(result.success)
        self.assertEqual(result.issues[0].code, "dataset_has_no_rows")

    def test_json_empty_array_is_invalid(self) -> None:
        storage = InMemoryStorage()
        storage.objects["workspace/.blobs/empty-json"] = b"[]"
        version = _build_version("dsv_empty_json", "json", "workspace/.blobs/empty-json")

        result = profile_dataset_version(version, storage)

        self.assertFalse(result.success)
        self.assertEqual(result.issues[0].code, "dataset_has_no_rows")

    def test_json_scalar_rows_are_invalid(self) -> None:
        storage = InMemoryStorage()
        storage.objects["workspace/.blobs/scalar-json"] = json.dumps([{"sku": "A"}, 123]).encode("utf-8")
        version = _build_version("dsv_scalar_json", "json", "workspace/.blobs/scalar-json")

        result = profile_dataset_version(version, storage)

        self.assertFalse(result.success)
        self.assertEqual(result.issues[0].code, "json_scalar_rows_not_supported")

    def test_xlsx_without_populated_sheets_is_invalid(self) -> None:
        workbook = Workbook()
        workbook.active.title = "Sales"
        payload = io.BytesIO()
        workbook.save(payload)
        storage = InMemoryStorage()
        storage.objects["workspace/.blobs/empty-xlsx"] = payload.getvalue()
        version = _build_version("dsv_empty_xlsx", "xlsx", "workspace/.blobs/empty-xlsx")

        result = profile_dataset_version(version, storage)

        self.assertFalse(result.success)
        self.assertEqual(result.issues[0].code, "dataset_has_no_rows")

    def test_damaged_xlsx_returns_invalid_issue(self) -> None:
        storage = InMemoryStorage()
        storage.objects["workspace/.blobs/bad-xlsx"] = b"not-a-workbook"
        version = _build_version("dsv_bad_xlsx", "xlsx", "workspace/.blobs/bad-xlsx")

        result = profile_dataset_version(version, storage)

        self.assertFalse(result.success)
        self.assertEqual(result.issues[0].code, "xlsx_invalid")

    def test_missing_storage_returns_invalid_issue(self) -> None:
        version = _build_version("dsv_missing", "csv", "workspace/.blobs/missing")

        result = profile_dataset_version(version, InMemoryStorage())

        self.assertFalse(result.success)
        self.assertEqual(result.issues[0].code, "storage_missing")

    def test_large_profile_sets_sampled_and_limits_examples(self) -> None:
        rows = ["id,value"]
        for index in range(10020):
            rows.append(f"{index},item-{index}")
        storage = InMemoryStorage()
        storage.objects["workspace/.blobs/large-csv"] = ("\n".join(rows) + "\n").encode("utf-8")
        version = _build_version("dsv_large_csv", "csv", "workspace/.blobs/large-csv")

        result = profile_dataset_version(version, storage)

        self.assertTrue(result.success)
        sheet = result.profile.sheets[0]
        self.assertTrue(sheet.sampled)
        self.assertEqual(len(sheet.columns[1].examples), 5)
        self.assertIsNone(sheet.columns[1].unique_count)


def _build_version(version_id: str, file_format: str, storage_key: str) -> DatasetVersion:
    return DatasetVersion(
        id=version_id,
        dataset_id="ds_1",
        original_filename=f"dataset.{file_format}",
        storage_key=storage_key,
        format=file_format,
        size_bytes=1,
        checksum_sha256=f"sum-{version_id}",
        status="profiling",
        created_at=datetime.now(UTC),
    )
