"""Dataset profiling for uploaded DatasetVersion objects."""

from __future__ import annotations

import csv
import io
import os
import json
from collections import defaultdict
from collections.abc import Iterable
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from application.datasets.jobs import DatasetProfileJobResult
from application.datasets.models import (
    DatasetIssue as AppDatasetIssue,
    DatasetColumnProfile,
    DatasetProfile as AppDatasetProfile,
    DatasetSheetProfile,
    DatasetVersion as AppDatasetVersion,
)
from application.datasets.repository import RawFileStorage

CSV_DELIMITERS = [",", ";", "\t", "|"]
CSV_ENCODINGS = [
    ("utf-8-sig", None),
    ("utf-8", None),
    ("cp1251", "CSV encoding cp1251 was detected"),
]
ROOT_SHEET_NAME = "__root__"
PROFILE_SAMPLE_ROWS = int(os.getenv("PROFILE_SAMPLE_ROWS", "10000"))
PROFILE_EXAMPLES_PER_COLUMN = int(os.getenv("PROFILE_EXAMPLES_PER_COLUMN", "5"))


class ProfileFailure(Exception):
    def __init__(self, code: str, message: str) -> None:
        super().__init__(message)
        self.code = code
        self.message = message


def profile_dataset_version(
    version: AppDatasetVersion,
    storage: RawFileStorage,
) -> DatasetProfileJobResult:
    if not version.storage_key or not version.format:
        return _invalid_result(
            version.id,
            "profile_metadata_missing",
            "Dataset version is missing storage metadata for profiling",
        )

    try:
        with storage.open_read(version.storage_key) as stream:
            if version.format == "csv":
                profile = _profile_uploaded_csv(stream)
            elif version.format == "json":
                profile = _profile_uploaded_json(stream)
            elif version.format == "xlsx":
                profile = _profile_uploaded_xlsx(stream)
            else:
                return _invalid_result(
                    version.id,
                    "unsupported_format",
                    f"Unsupported dataset format: {version.format}",
                )
    except ProfileFailure as exc:
        return _invalid_result(version.id, exc.code, exc.message)
    except FileNotFoundError:
        return _invalid_result(
            version.id,
            "storage_missing",
            "Stored dataset file is missing",
        )
    except UnicodeDecodeError:
        return _invalid_result(
            version.id,
            "csv_encoding_unsupported",
            "CSV encoding is unsupported",
        )
    except json.JSONDecodeError:
        return _invalid_result(
            version.id,
            "json_invalid",
            "JSON payload could not be parsed",
        )
    except Exception:
        return _invalid_result(
            version.id,
            "profile_runtime_error",
            "Dataset profiling failed unexpectedly",
        )

    return DatasetProfileJobResult(
        version_id=version.id,
        profile=profile,
        issues=[],
        success=True,
    )


def _profile_uploaded_csv(stream) -> AppDatasetProfile:
    warnings: list[str] = []
    sample_bytes = stream.read(4096)
    stream.seek(0)
    chosen_encoding = None
    sample_text = None

    for encoding, warning in CSV_ENCODINGS:
        try:
            sample_text = sample_bytes.decode(encoding)
            chosen_encoding = encoding
            if warning:
                warnings.append(warning)
            break
        except UnicodeDecodeError:
            continue

    if chosen_encoding is None or sample_text is None:
        raise UnicodeDecodeError("csv", sample_bytes, 0, len(sample_bytes), "unsupported encoding")

    delimiter = ","
    if sample_text.strip():
        try:
            dialect = csv.Sniffer().sniff(sample_text, delimiters="".join(CSV_DELIMITERS))
            delimiter = dialect.delimiter
            if delimiter != ",":
                warnings.append(f"CSV delimiter '{delimiter}' was detected")
        except csv.Error:
            warnings.append("CSV delimiter detection fell back to comma")

    stream.seek(0)
    try:
        text_stream = io.TextIOWrapper(stream, encoding=chosen_encoding, newline="")
        reader = csv.reader(text_stream, delimiter=delimiter)
        header = None
        rows: list[list[Any]] = []
        total_rows = 0
        sampled = False

        for row in reader:
            if not any(_has_value(cell) for cell in row):
                continue
            if header is None:
                header = _normalize_header(row, warnings)
                continue
            total_rows += 1
            if len(rows) < PROFILE_SAMPLE_ROWS:
                rows.append(row)
            else:
                sampled = True
    except UnicodeDecodeError as exc:
        raise exc
    finally:
        try:
            text_stream.detach()
        except Exception:
            pass

    if header is None:
        raise ProfileFailure("csv_invalid", "CSV file is empty")

    if total_rows == 0:
        raise ProfileFailure("dataset_has_no_rows", "Dataset has no data rows")

    return AppDatasetProfile(
        format="csv",
        sheets=[
            _build_sheet_profile(
                ROOT_SHEET_NAME,
                header,
                rows,
                total_rows,
                warnings,
                sampled,
            )
        ],
        warnings=warnings,
    )


def _profile_uploaded_json(stream) -> AppDatasetProfile:
    warnings: list[str] = []
    with io.TextIOWrapper(stream, encoding="utf-8-sig") as text_stream:
        parser = _JsonStreamParser(text_stream)
        payload_type = parser.peek_container_type()
        if payload_type == "object":
            item = parser.read_single_object()
            rows = [item]
            total_rows = 1
            sampled = False
        elif payload_type == "array":
            rows = []
            total_rows = 0
            sampled = False
            for item in parser.iter_array_items():
                if not isinstance(item, dict):
                    raise ProfileFailure(
                        "json_scalar_rows_not_supported",
                        "JSON array must contain only object rows",
                    )
                total_rows += 1
                if len(rows) < PROFILE_SAMPLE_ROWS:
                    rows.append(item)
                else:
                    sampled = True

            if total_rows == 0:
                raise ProfileFailure("dataset_has_no_rows", "Dataset has no data rows")
        else:
            raise ProfileFailure(
                "json_invalid",
                "JSON root must be an object or an array of objects",
            )

    header = _collect_json_header(rows)

    return AppDatasetProfile(
        format="json",
        sheets=[
            _build_sheet_profile(
                ROOT_SHEET_NAME,
                header,
                rows,
                total_rows,
                warnings,
                sampled,
            )
        ],
        warnings=warnings,
    )


def _profile_uploaded_xlsx(stream) -> AppDatasetProfile:
    warnings: list[str] = []
    try:
        workbook = load_workbook(stream, read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
        raise ProfileFailure("xlsx_invalid", "Workbook is invalid") from exc

    sheets: list[DatasetSheetProfile] = []
    populated_sheets = 0

    try:
        for worksheet in workbook.worksheets:
            sheet_warnings: list[str] = []
            header = None
            rows: list[list[Any]] = []
            total_rows = 0
            sampled = False
            for row in worksheet.iter_rows(values_only=True):
                values = list(row)
                if not any(_has_value(cell) for cell in values):
                    continue
                if header is None:
                    header = _normalize_header(values, sheet_warnings)
                    continue
                total_rows += 1
                if len(rows) < PROFILE_SAMPLE_ROWS:
                    rows.append(values)
                else:
                    sampled = True

            if header is None:
                warnings.append(f"Worksheet '{worksheet.title}' is empty")
                continue

            if total_rows == 0:
                warnings.append(f"Worksheet '{worksheet.title}' has no data rows")
                sheets.append(
                    DatasetSheetProfile(
                        name=worksheet.title,
                        row_count=0,
                        columns=[],
                        warnings=["dataset_has_no_rows"],
                        sampled=sampled,
                    )
                )
                continue

            populated_sheets += 1
            sheets.append(
                _build_sheet_profile(
                    worksheet.title,
                    header,
                    rows,
                    total_rows,
                    sheet_warnings,
                    sampled,
                )
            )

        if populated_sheets == 0:
            raise ProfileFailure("dataset_has_no_rows", "Workbook has no populated worksheets")
    finally:
        workbook.close()

    return AppDatasetProfile(
        format="xlsx",
        sheets=sheets,
        warnings=warnings,
    )


def _build_sheet_profile(
    name: str,
    header: list[str],
    rows: list[dict[str, Any] | list[Any]],
    total_rows: int,
    warnings: list[str],
    sampled: bool,
) -> DatasetSheetProfile:
    row_mappings = [_row_to_mapping(header, row) for row in rows]
    columns = [
        _build_column_profile(column_name, [mapping.get(column_name) for mapping in row_mappings], total_rows)
        for column_name in header
    ]
    return DatasetSheetProfile(
        name=name,
        row_count=total_rows,
        columns=columns,
        warnings=warnings,
        sampled=sampled,
    )


def _build_column_profile(
    name: str,
    values: list[Any],
    total_rows: int,
) -> DatasetColumnProfile:
    non_null_values = [value for value in values if _has_value(value)]
    inferred_type = _infer_column_type(non_null_values)
    examples = _collect_examples(non_null_values)
    unique_count = None if total_rows > PROFILE_SAMPLE_ROWS else len({_stable_value_key(value) for value in non_null_values})
    min_value, max_value = _compute_range(non_null_values, inferred_type)
    null_ratio = 1.0 if total_rows == 0 else (total_rows - len(non_null_values)) / total_rows
    return DatasetColumnProfile(
        name=name,
        inferred_type=inferred_type,
        null_ratio=null_ratio,
        unique_count=unique_count,
        examples=examples,
        min_value=min_value,
        max_value=max_value,
    )


def _normalize_header(values: list[Any], warnings: list[str]) -> list[str]:
    header: list[str] = []
    for index, value in enumerate(values, start=1):
        if _has_value(value):
            header.append(str(value).strip())
        else:
            header.append(f"column_{index}")
            warnings.append(f"Header column {index} was blank and was renamed")
    return header


def _has_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return value.strip() != ""
    return True


def _row_to_mapping(header: list[str], row: dict[str, Any] | list[Any]) -> dict[str, Any]:
    if isinstance(row, dict):
        return {column_name: row.get(column_name) for column_name in header}

    return {
        column_name: row[index] if index < len(row) else None
        for index, column_name in enumerate(header)
    }


def _collect_json_header(rows: list[dict[str, Any]]) -> list[str]:
    header: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for key in row.keys():
            key_name = str(key)
            if key_name not in seen:
                seen.add(key_name)
                header.append(key_name)
    return header


def _infer_column_type(values: list[Any]) -> str:
    if not values:
        return "unknown"

    detectors = [
        ("boolean", _is_boolean),
        ("integer", _is_integer),
        ("number", _is_number),
        ("datetime", _is_datetime),
        ("string", lambda _: True),
    ]
    for inferred_type, detector in detectors:
        if all(detector(value) for value in values):
            return inferred_type
    return "unknown"


def _collect_examples(values: list[Any]) -> list[str]:
    examples: list[str] = []
    seen: set[str] = set()
    for value in values:
        rendered = _stringify_value(value)
        if rendered not in seen:
            seen.add(rendered)
            examples.append(rendered)
        if len(examples) >= PROFILE_EXAMPLES_PER_COLUMN:
            break
    return examples


def _compute_range(values: list[Any], inferred_type: str) -> tuple[str | None, str | None]:
    if not values or inferred_type not in {"integer", "number", "datetime"}:
        return None, None

    normalized: list[Any] = []
    for value in values:
        converted = _normalize_for_range(value, inferred_type)
        if converted is None:
            return None, None
        normalized.append(converted)

    return _stringify_value(min(normalized)), _stringify_value(max(normalized))


def _normalize_for_range(value: Any, inferred_type: str):
    if inferred_type == "datetime":
        if isinstance(value, datetime):
            return value
        if isinstance(value, str):
            try:
                return datetime.fromisoformat(value)
            except ValueError:
                return None
        return None

    if inferred_type == "integer":
        try:
            return int(str(value).strip())
        except (TypeError, ValueError):
            return None

    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, AttributeError):
        return None


def _is_boolean(value: Any) -> bool:
    if isinstance(value, bool):
        return True
    if isinstance(value, str):
        return value.strip().lower() in {"true", "false", "yes", "no", "0", "1"}
    return False


def _is_integer(value: Any) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, int):
        return True
    if isinstance(value, str):
        try:
            int(value.strip())
            return True
        except ValueError:
            return False
    return False


def _is_number(value: Any) -> bool:
    if _is_integer(value):
        return True
    if isinstance(value, bool):
        return False
    if isinstance(value, float):
        return True
    if isinstance(value, str):
        try:
            Decimal(value.strip())
            return True
        except InvalidOperation:
            return False
    return False


def _is_datetime(value: Any) -> bool:
    if isinstance(value, datetime):
        return True
    if isinstance(value, str):
        try:
            datetime.fromisoformat(value.strip())
            return True
        except ValueError:
            return False
    return False


def _stringify_value(value: Any) -> str:
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value)


def _stable_value_key(value: Any) -> str:
    return f"{type(value).__name__}:{_stringify_value(value)}"


class _JsonStreamParser:
    def __init__(self, text_stream) -> None:
        self._text_stream = text_stream
        self._decoder = json.JSONDecoder()
        self._buffer = ""
        self._position = 0
        self._eof = False

    def peek_container_type(self) -> str:
        token = self._next_nonspace_char()
        if token is None:
            raise ProfileFailure("json_invalid", "JSON payload is empty")
        self._position -= 1
        if token == "{":
            return "object"
        if token == "[":
            return "array"
        return "scalar"

    def read_single_object(self) -> dict[str, Any]:
        value = self._decode_next()
        if not isinstance(value, dict):
            raise ProfileFailure("json_invalid", "JSON object root is invalid")
        trailing = self._next_nonspace_char()
        if trailing is not None:
            raise json.JSONDecodeError("Trailing data", self._buffer, self._position)
        return value

    def iter_array_items(self) -> Iterable[Any]:
        opening = self._next_nonspace_char()
        if opening != "[":
            raise json.JSONDecodeError("Expected JSON array", self._buffer, self._position)

        token = self._next_nonspace_char()
        if token == "]":
            return
        if token is None:
            raise json.JSONDecodeError("Unexpected EOF", self._buffer, self._position)
        self._position -= 1

        while True:
            yield self._decode_next()
            separator = self._next_nonspace_char()
            if separator == "]":
                return
            if separator != ",":
                raise json.JSONDecodeError("Expected ',' or ']'", self._buffer, self._position)

    def _decode_next(self) -> Any:
        while True:
            self._consume_whitespace()
            try:
                value, end = self._decoder.raw_decode(self._buffer, self._position)
                self._position = end
                return value
            except json.JSONDecodeError:
                if not self._fill():
                    raise

    def _next_nonspace_char(self) -> str | None:
        while True:
            self._consume_whitespace()
            if self._position < len(self._buffer):
                value = self._buffer[self._position]
                self._position += 1
                return value
            if not self._fill():
                return None

    def _consume_whitespace(self) -> None:
        while True:
            while self._position < len(self._buffer) and self._buffer[self._position].isspace():
                self._position += 1
            if self._position < len(self._buffer) or not self._fill():
                return

    def _fill(self) -> bool:
        if self._eof:
            return False
        chunk = self._text_stream.read(4096)
        if chunk == "":
            self._eof = True
            return False
        if self._position > 0:
            self._buffer = self._buffer[self._position :]
            self._position = 0
        self._buffer += chunk
        return True


def _invalid_result(version_id: str, code: str, message: str) -> DatasetProfileJobResult:
    return DatasetProfileJobResult(
        version_id=version_id,
        profile=None,
        issues=[AppDatasetIssue(code=code, message=message, severity="error")],
        success=False,
    )
